import time
from openpyxl import load_workbook

import Train
import SensorNode
import Sensor

class BigScheduleTable(object):
   def __init__(self):
      self.scheduleTable = dict()
      self.stationIds = list()
      self.subwayIds = list()
      self.str_startTime = ""
      self.str_stopTime = ""
      self.numOfSubways = 0
      self.numOfTimes = 0
      self.numOfStations = 0

   def loadYolpon2Pyonsong(self, filename):
      wb = load_workbook(filename)
      sheet = wb.worksheets[0]
      yolpon2pyonsong = dict()
	  
      for i in range(5,31):
         yolpon2pyonsong[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=3).value
         #print sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value
      #for el in yolpon2pyonsong:
      #   print el," => ",yolpon2pyonsong[el]
      return yolpon2pyonsong

   def loadFromCSV(self, filename):
      fi = open(filename, 'r')
      self.str_startTime, self.str_stopTime = fi.readline().strip().split(',')
      #print 'startTime:', str_startTime, 'stopTime:', str_stopTime

      self.numOfSubways, self.numOfTimes = map(int, fi.readline().strip().split(','))

      self.stationIds = map(int, fi.readline().strip().split(','))
      self.numOfStations = len(self.stationIds)

      _subwaysAr = fi.readline().strip().split(',')
      self.subwayIds = map(int, _subwaysAr[1:])

      

      for _ in range(self.numOfTimes):
         rline = fi.readline()
         #print 'at for cycle:', rline
         self.parseTimeLine(rline)
      fi.close()
   

   def parseTrainDirection(self, inpstr):
      #print 'inpstr at parseTrainDirection :', inpstr
      strTrainNum, _dir = inpstr.strip().split('_')
      intTrainNum = int(strTrainNum)
      dir = 1 if _dir=='d' else 2
      return intTrainNum, dir


   def parseTimeLine(self, _inpstr): #, scheduleTable, subwaysAr):
      inpstr = _inpstr.strip().split(',')
      sub_time = inpstr[0]
      trains = list()
      for index in range(1,len(inpstr)):
         #print 'inpstr[index]:', inpstr[index]
         if inpstr[index] != '0':
            intTrainNum, intDir = self.parseTrainDirection(inpstr[index])
            trains.append([self.subwayIds[index-1], intTrainNum, intDir])
      self.scheduleTable[sub_time] = trains

   def requestCurrentStatus(self, timeInpStr):

      if not self.scheduleTable:
         print 'No schedule table has been loaded into memory'
         return None

      try:
         #check timeInpStr being time:
         hour, minute = map(int, timeInpStr.strip().split(':'))
         if timeInpStr in self.scheduleTable:
            return self.scheduleTable[timeInpStr]
         else:
            print 'Schedule does not contain this kind of number'
            return None

      except ValueError:
         print 'Oops! , parsing the time throw an error: ', timeInpStr
         return None

      #return scheduleTable[timeInpStr]

   #returns Train[] array which contains the train object initialized with dummy sensors
   def initializeTrainTables(self,TrainNum2ID, NumberOfSensors):
      Trains = dict()
      for trainID in self.subwayIds:
         subwayNum = trainID
         subwayID = TrainNum2ID[subwayNum]
         Trains[subwayID] = Train.Train(subwayID)

         """  DO not create any sensors for the New Train
         #adding NUMBER_OF_SENSORS dummy sensor nodes
         temporary_temp_int = 0 #int(time.strftime("%M"))
         for i in range(1, NumberOfSensors + 1):
            curTrain_DummySensorNode = SensorNode.SensorNode(i, 'TempHum', False)
            curTrain_DummySensorNode.sensors['temp'] = Sensor.Sensor('temp', 'celcius', temporary_temp_int)
            curTrain_DummySensorNode.sensors['hum'] = Sensor.Sensor('hum', '%', temporary_temp_int)
            Trains[subwayNum].sensorNodes[i] = curTrain_DummySensorNode
         """
      return Trains

""" Example code 

bigtable = BigScheduleTable()
bigtable.loadFromCSV('schedule.csv')

print 'Enter the time:',
curtime = raw_input()
print bigtable.requestCurrentStatus(curtime)
print bigtable.loadYolpon2Pyonsong("excel1.xlsx")


# result of above example call:
#
# Enter the time: 15:03
# [[1905, 118, 2], [1006, 107, 2], [1919, 114, 1], [1923, 102, 1], [1925, 101, 2], [1927, 108, 1], [1929, 119, 1], [1931, 113, 2]]



print time.time()
print time.strftime("%H:%M")
"""



""" Non Class Code 
fi = open('schedule.csv', 'r')


str_startTime, str_stopTime = fi.readline().strip().split(',')
#print 'startTime:', str_startTime, 'stopTime:', str_stopTime

numOfSubways, numOfTimes = map(int,fi.readline().strip().split(','))

stationAr = map(int, fi.readline().strip().split(','))
_subwaysAr = fi.readline().strip().split(',')
subwaysAr = map(int, _subwaysAr[1:])

scheduleTable = dict()
for _ in range(numOfTimes):
   rline = fi.readline()
   #print 'at for cycle:', rline
   parseTimeLine(rline, scheduleTable, subwaysAr)

#print 'ANSWER'
#print scheduleTable

# query section
print 'Enter the time:',
curtime = raw_input()
print requestCurrentStatus(curtime, scheduleTable)
"""





