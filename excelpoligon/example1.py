#!/usr/bin/env python
# -*- coding: utf-8 -*- 

from openpyxl import load_workbook

wb = load_workbook("./excel1.xlsx")

print wb.get_sheet_names()

#sheet = wb.get_sheet_by_name('평일') #  'Sheet1')
sheet = wb.worksheets[0]

print sheet.title

print '-------------'
#print sheet['B5'].value
for i in range(1,16):
   for j in range(1,40):
      if sheet.cell(row=i, column=j).value != None:
         print sheet.cell(row=i, column=j).value,
      else:
         print ' ',
   print '\n'


print '------ this is what we need -------'
for i in range(5,31):
   print sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value

#anotherSheet = wb.active
#print anotherSheet
