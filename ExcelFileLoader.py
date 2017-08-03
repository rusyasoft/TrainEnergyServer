#!/usr/bin/env python
# -*- coding: utf-8 -*- 

from openpyxl import load_workbook

class ExcelFileLoader(object):
	def __init__(self):
		self.trainIDcolumn = None
		self.trainIDDict = None
        #	self.fullfilepath = fullfilepath

	def loadfile(self, fullfilepath):
		self.fullfilepath = fullfilepath
		self.wb = load_workbook(fullfilepath)

		#print wb.get_sheet_names()

		#sheet = wb.get_sheet_by_name('평일') #  'Sheet1')
		self.sheet = self.wb.worksheets[0]

		#print sheet.title

		#print '-------------'
        #print sheet['B5'].value
		self.trainIDcolumn = list()
		print '------ this is what we need -------'
		for i in range(5,31):
			#print sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value
			self.trainIDcolumn.append( (self.sheet.cell(row=i, column=2).value, self.sheet.cell(row=i, column=3).value) )
		#return self.trainIDcolumn
		#converting list to dictionary
		self.trainNum2ID = dict()
		for val1, val2 in self.trainIDcolumn:
			self.trainNum2ID[int(val1)] = int(val2)
		#print "result of convertion from list of trains to dicitionary ->", trainIDDict
		return self.trainNum2ID

    

	def printTrainList(self):
		for l in self.trainIDcolumn:
			print l

